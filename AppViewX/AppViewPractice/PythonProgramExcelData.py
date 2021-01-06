from openpyxl import Workbook
from openpyxl import load_workbook

with open("textFile.txt", 'w') as write:
    write.write("Testing .txt file data for AppViewX")

def excel_data():
    with open("textFile.txt", 'r') as reader:
        print(reader.read())

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Gururaj'
    ws['B1'] = '9880370312'
    ws['A2'] = 'Kashinath'
    ws['B2'] = '9591376266'

    wb.save("Sample_File.xlsx")

    load_wb = load_workbook("Sample_File.xlsx")
    sheet = load_wb.active
    print(sheet['A1'].value)
    print(sheet['A2'].value)
    print(sheet['B1'].value)
    print(sheet['B2'].value)

    print(sheet.cell(row=1, column=1).value)





